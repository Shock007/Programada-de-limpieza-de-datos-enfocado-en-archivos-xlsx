import tkinter as tk
from tkinterdnd2 import DND_FILES, TkinterDnD
from tkinter import messagebox, filedialog
from tkinter import ttk
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ──────────────────────────────────────────────────────────────────────────────
#  Paleta de colores
# ──────────────────────────────────────────────────────────────────────────────
BG_BASE      = "#F5F5F5"   # fondo general de la ventana
BG_CARD      = "#FFFFFF"   # fondo de paneles / tarjetas
BG_ALT_ROW   = "#EBF5FB"   # fila par en la tabla (azul muy suave)

CLR_GREEN    = "#27AE60"   # botones primarios / éxito
CLR_GREEN_H  = "#1E8449"   # hover verde
CLR_RED      = "#E74C3C"   # botones destructivos / alertas
CLR_RED_H    = "#C0392B"   # hover rojo
CLR_BLUE     = "#2980B9"   # botones de análisis
CLR_BLUE_H   = "#1A5276"   # hover azul
CLR_ORANGE   = "#E67E22"   # advertencias

ST_IDLE      = "#7F8C8D"   # barra de estado — en espera
ST_OK        = "#1E8449"   # barra de estado — éxito
ST_ERROR     = "#C0392B"   # barra de estado — error
ST_WARN      = "#E67E22"   # barra de estado — advertencia / hallazgo

# ──────────────────────────────────────────────────────────────────────────────
#  Constantes de dominio
# ──────────────────────────────────────────────────────────────────────────────
NUMBER_FORMAT_MAP = {
    "General":               "General",
    "Número":                "#,##0.00",
    "Moneda Colombiana":     "[$$-2058]#,##0.00;[RED]-[$$-2058]#,##0.00",
    "Moneda Estadounidense": "$#,##0.00",
    "Fecha":                 "dd/mm/yyyy",
    "Hora":                  "hh:mm:ss",
    "Porcentaje":            "0.00%",
    "Texto":                 "@",
}

FONTS_AVAILABLE   = ["Arial", "Calibri"]
ALIGN_OPTIONS     = ["left", "center", "right"]
VALIGN_OPTIONS    = ["top", "center", "bottom"]
FONT_SIZE_DEFAULT = 11
FONT_SIZE_MIN     = 8
FONT_SIZE_MAX     = 72

COLOR_CELL_SELECT = "#AED6F1"   # azul — celda seleccionada
COLOR_COL_SELECT  = "#A9DFBF"   # verde — columna seleccionada


# ──────────────────────────────────────────────────────────────────────────────
#  Utilidad: botón con color (tk.Button se colorea mejor que ttk en Windows)
# ──────────────────────────────────────────────────────────────────────────────
def _colored_btn(parent, text: str, command, bg: str, hover: str,
                 state: str = tk.NORMAL) -> tk.Button:
    """Crea un tk.Button con fondo de color y efecto hover."""
    btn = tk.Button(
        parent, text=text, command=command,
        bg=bg, fg="white", activebackground=hover, activeforeground="white",
        relief=tk.FLAT, bd=0, padx=8, pady=5,
        font=("TkDefaultFont", 9, "bold"),
        cursor="hand2", state=state,
        disabledforeground="#CCCCCC",
    )
    btn.bind("<Enter>", lambda _: btn.config(bg=hover) if btn["state"] != tk.DISABLED else None)
    btn.bind("<Leave>", lambda _: btn.config(bg=bg)    if btn["state"] != tk.DISABLED else None)
    return btn


class DataCleanerApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("XLSX Data Cleaner")
        self.root.geometry("1250x760")
        self.root.configure(bg=BG_BASE)

        # ── Estado interno ──────────────────────────────────────────────────
        self.current_file:         str | None          = None
        self.dataframe:            pd.DataFrame | None = None
        self.edited_cells:         dict                = {}
        self.cell_formats:         dict                = {}
        self.selected_cell:        tuple | None        = None
        self.selected_column:      int   | None        = None
        self.row_id_map:           dict                = {}
        self._exact_dup_col_pairs: list                = []

        self._status_after_id = None   # ID del after() de limpieza de status

        self._setup_styles()
        self._build_ui()

    # ══════════════════════════════════════════════════════════════════════════
    #  ESTILOS ttk
    # ══════════════════════════════════════════════════════════════════════════

    def _setup_styles(self) -> None:
        style = ttk.Style()
        style.theme_use("clam")

        # Contenedores
        style.configure("TFrame",          background=BG_BASE)
        style.configure("Card.TFrame",     background=BG_CARD)
        style.configure("TLabelframe",     background=BG_CARD, bordercolor="#D5D8DC")
        style.configure("TLabelframe.Label",
                        background=BG_CARD, foreground="#2C3E50",
                        font=("TkDefaultFont", 9, "bold"))

        # Notebook
        style.configure("TNotebook",       background=BG_BASE, borderwidth=0)
        style.configure("TNotebook.Tab",   background="#D5D8DC", padding=[14, 5],
                        font=("TkDefaultFont", 9))
        style.map("TNotebook.Tab",
                  background=[("selected", BG_CARD)],
                  foreground=[("selected", "#2C3E50")])

        # Labels, Radiobuttons, Separators
        style.configure("TLabel",       background=BG_CARD,  foreground="#2C3E50")
        style.configure("Toolbar.TLabel", background=BG_BASE, foreground="#555555")
        style.configure("TRadiobutton", background=BG_CARD,  foreground="#2C3E50")
        style.configure("TSeparator",   background="#D5D8DC")

        # Treeview
        style.configure("Treeview",
                        background=BG_CARD, fieldbackground=BG_CARD,
                        foreground="#2C3E50", rowheight=22,
                        font=("TkDefaultFont", 9))
        style.configure("Treeview.Heading",
                        background="#2C3E50", foreground="white",
                        font=("TkDefaultFont", 9, "bold"), relief=tk.FLAT)
        style.map("Treeview.Heading",
                  background=[("active", CLR_BLUE)])
        style.map("Treeview",
                  background=[("selected", COLOR_CELL_SELECT)],
                  foreground=[("selected", "#1A252F")])

        # Scrollbars
        style.configure("TScrollbar", background="#CCCCCC", troughcolor=BG_BASE,
                        arrowcolor="#888888")

        # Entradas, Combobox, Spinbox
        style.configure("TEntry",    fieldbackground="white", foreground="#2C3E50")
        style.configure("TCombobox", fieldbackground="white", foreground="#2C3E50")
        style.configure("TSpinbox",  fieldbackground="white", foreground="#2C3E50")

    # ══════════════════════════════════════════════════════════════════════════
    #  CONSTRUCCIÓN DE LA INTERFAZ
    # ══════════════════════════════════════════════════════════════════════════

    def _build_ui(self) -> None:
        # Barra de estado al fondo (se empaca primero para quedar abajo)
        self._build_status_bar()

        main = ttk.Frame(self.root)
        main.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 4))
        self._build_toolbar(main)
        self._build_content(main)

    # ── Barra de estado ───────────────────────────────────────────────────────

    def _build_status_bar(self) -> None:
        self.status_bar = tk.Label(
            self.root, text="  ⏳ En espera — carga un archivo para comenzar.",
            bg=ST_IDLE, fg="white", anchor=tk.W,
            padx=10, pady=5, font=("TkDefaultFont", 9))
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def _set_status(self, msg: str, level: str = "ok", autoclean: bool = True) -> None:
        """
        Actualiza la barra de estado.
        level: "ok" | "error" | "warn" | "idle"
        """
        colors = {"ok": ST_OK, "error": ST_ERROR, "warn": ST_WARN, "idle": ST_IDLE}
        icons  = {"ok": "✔", "error": "✖", "warn": "⚠", "idle": "⏳"}
        bg = colors.get(level, ST_IDLE)
        ic = icons.get(level, "•")

        self.status_bar.config(bg=bg, text=f"  {ic}  {msg}")

        # Cancelar limpieza anterior si la hay
        if self._status_after_id:
            self.root.after_cancel(self._status_after_id)
            self._status_after_id = None

        if autoclean and level in ("ok", "warn"):
            self._status_after_id = self.root.after(
                6000,
                lambda: self.status_bar.config(
                    bg=ST_IDLE,
                    text="  ⏳ En espera."))

    # ── Barra de herramientas ─────────────────────────────────────────────────

    def _build_toolbar(self, parent: ttk.Frame) -> None:
        bar = tk.Frame(parent, bg=BG_BASE)
        bar.pack(fill=tk.X, pady=(0, 10))

        self.upload_btn = _colored_btn(
            bar, "📥 Cargar archivo", self.upload_file, CLR_GREEN, CLR_GREEN_H)
        self.upload_btn.pack(side=tk.LEFT, padx=(0, 6))

        self.save_btn = _colored_btn(
            bar, "💾 Guardar cambios", self.save_file,
            CLR_BLUE, CLR_BLUE_H, state=tk.DISABLED)
        self.save_btn.pack(side=tk.LEFT, padx=(0, 12))

        # Separador visual
        tk.Frame(bar, width=2, bg="#D5D8DC").pack(side=tk.LEFT, fill=tk.Y, padx=4)

        self.file_label = ttk.Label(bar, text="Ningún archivo cargado",
                                    style="Toolbar.TLabel")
        self.file_label.pack(side=tk.LEFT, padx=8)
                # ─────────────────────────────────────────────
        # Zona Drag & Drop
        # ─────────────────────────────────────────────
        self.drop_label = tk.Label(
            bar,
            text="📂 Arrastra un archivo XLSX aquí",
            bg="#D6EAF8",
            fg="#1B4F72",
            relief=tk.GROOVE,
            bd=2,
            padx=15,
            pady=8,
            font=("TkDefaultFont", 9, "bold")
        )

        self.drop_label.pack(side=tk.RIGHT, padx=10)

        # Registrar zona de arrastre
        self.drop_label.drop_target_register(DND_FILES)
        self.drop_label.dnd_bind("<<Drop>>", self._drop_file)

    # ── Área principal ────────────────────────────────────────────────────────

    def _build_content(self, parent: ttk.Frame) -> None:
        content = ttk.Frame(parent)
        content.pack(fill=tk.BOTH, expand=True)
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=1)

        self._build_table(content)

        self.notebook = ttk.Notebook(content)
        self.notebook.grid(row=0, column=1, sticky=tk.NSEW, padx=(6, 0))

        tab_edit = ttk.Frame(self.notebook, style="Card.TFrame")
        self.notebook.add(tab_edit, text="  ✏  Editar  ")
        self._build_edit_panel(tab_edit)

        tab_auto = ttk.Frame(self.notebook, style="Card.TFrame")
        self.notebook.add(tab_auto, text="  ⚙  Automatización  ")
        self._build_auto_panel(tab_auto)

    # ── Tabla Treeview ────────────────────────────────────────────────────────

    def _build_table(self, parent: ttk.Frame) -> None:
        frame = ttk.Frame(parent)
        frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(0, 4))
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL)

        self.tree = ttk.Treeview(
            frame, yscrollcommand=vsb.set,
            xscrollcommand=hsb.set, selectmode="browse"
        )
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        self.tree.grid(row=0, column=0, sticky=tk.NSEW)
        vsb.grid(row=0, column=1, sticky=tk.NS)
        hsb.grid(row=1, column=0, sticky=tk.EW)

        # Tags de resaltado
        self.tree.tag_configure("cell_selected", background=COLOR_CELL_SELECT)
        self.tree.tag_configure("col_selected",  background=COLOR_COL_SELECT)
        self.tree.tag_configure("oddrow",        background=BG_CARD)
        self.tree.tag_configure("evenrow",       background=BG_ALT_ROW)

        self.tree.bind("<Button-1>", self._on_click)
        self.tree.bind("<Double-1>", self._on_double_click)

    # ── Panel "Editar" ────────────────────────────────────────────────────────

    def _build_edit_panel(self, parent: ttk.Frame) -> None:
        panel = ttk.LabelFrame(parent, text="Editar", padding=14)
        panel.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        # Info de celda seleccionada
        self.cell_info_label = ttk.Label(
            panel, text="Ninguna celda seleccionada",
            relief=tk.SUNKEN, anchor=tk.W, padding=(6, 4))
        self.cell_info_label.pack(fill=tk.X, pady=(0, 12))

        # Valor
        self._section_label(panel, "Valor de la celda")
        self.value_entry = ttk.Entry(panel, width=22)
        self.value_entry.pack(fill=tk.X, pady=(0, 12))

        ttk.Separator(panel, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)

        # Fuente
        self._section_label(panel, "Fuente")
        self.font_var = tk.StringVar(value="Calibri")
        ttk.Combobox(panel, textvariable=self.font_var, width=20,
                     values=FONTS_AVAILABLE, state="readonly").pack(fill=tk.X, pady=(0, 10))

        # Tamaño
        self._section_label(panel, "Tamaño de fuente")
        self.font_size_var = tk.StringVar(value=str(FONT_SIZE_DEFAULT))
        ttk.Spinbox(panel, from_=FONT_SIZE_MIN, to=FONT_SIZE_MAX,
                    textvariable=self.font_size_var,
                    width=20).pack(fill=tk.X, pady=(0, 10))

        ttk.Separator(panel, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)

        # Alineación horizontal
        self._section_label(panel, "Alineación horizontal")
        self.h_align_var = tk.StringVar(value="left")
        frm_h = ttk.Frame(panel)
        frm_h.pack(fill=tk.X, pady=(0, 10))
        for lbl, val in [("← Izquierda", "left"),
                          ("↔ Centro",    "center"),
                          ("→ Derecha",   "right")]:
            ttk.Radiobutton(frm_h, text=lbl, variable=self.h_align_var,
                            value=val).pack(anchor=tk.W)

        # Alineación vertical
        self._section_label(panel, "Alineación vertical")
        self.v_align_var = tk.StringVar(value="center")
        frm_v = ttk.Frame(panel)
        frm_v.pack(fill=tk.X, pady=(0, 10))
        for lbl, val in [("↑ Superior", "top"),
                          ("↕ Medio",   "center"),
                          ("↓ Inferior","bottom")]:
            ttk.Radiobutton(frm_v, text=lbl, variable=self.v_align_var,
                            value=val).pack(anchor=tk.W)

        ttk.Separator(panel, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)

        # Formato de número
        self._section_label(panel, "Formato de número")
        self.number_format_var = tk.StringVar(value="General")
        ttk.Combobox(panel, textvariable=self.number_format_var, width=20,
                     values=list(NUMBER_FORMAT_MAP.keys()),
                     state="readonly").pack(fill=tk.X, pady=(0, 14))

        # Botón aplicar (verde)
        self.apply_btn = _colored_btn(
            panel, "✔  Aplicar formato", self._apply_formatting,
            CLR_GREEN, CLR_GREEN_H, state=tk.DISABLED)
        self.apply_btn.pack(fill=tk.X, ipady=2)

    @staticmethod
    def _section_label(parent, text: str) -> None:
        ttk.Label(parent, text=text,
                  font=("TkDefaultFont", 8, "bold"),
                  foreground="#2C3E50").pack(anchor=tk.W, pady=(4, 3))

    # ══════════════════════════════════════════════════════════════════════════
    #  PESTAÑA AUTOMATIZACIÓN
    # ══════════════════════════════════════════════════════════════════════════

    def _build_auto_panel(self, parent: ttk.Frame) -> None:
        # Canvas + Scrollbar para contenido largo
        canvas = tk.Canvas(parent, bg=BG_BASE, borderwidth=0, highlightthickness=0)
        sb = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        inner = ttk.Frame(canvas)
        win_id = canvas.create_window((0, 0), window=inner, anchor=tk.NW)

        canvas.bind("<Configure>",        lambda e: canvas.itemconfig(win_id, width=e.width))
        inner.bind("<Configure>",         lambda _: canvas.configure(
                                              scrollregion=canvas.bbox("all")))

        # ── Botón único de escaneo ────────────────────────────────────────────
        scan_frame = tk.Frame(inner, bg=BG_BASE)
        scan_frame.pack(fill=tk.X, padx=8, pady=(10, 6))

        _colored_btn(
            scan_frame, "🔍  Escanear archivo completo",
            self._scan_all, CLR_BLUE, CLR_BLUE_H
        ).pack(fill=tk.X, ipady=4)

        ttk.Separator(inner, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=8, pady=(4, 0))

        # ════════════════════════════════════════════════════════════════════
        # SECCIÓN 1 — VACÍOS
        # ════════════════════════════════════════════════════════════════════
        sec_e = ttk.LabelFrame(inner, text="  Celdas y columnas vacías  ", padding=10)
        sec_e.pack(fill=tk.X, padx=8, pady=(10, 4))

        self._section_label(sec_e, "Resumen general")
        self.lbl_total_empty_cells = ttk.Label(sec_e, text="Celdas vacías: —",
            foreground=CLR_RED, font=("TkDefaultFont", 9, "bold"))
        self.lbl_total_empty_cells.pack(anchor=tk.W, pady=1)

        self.lbl_total_empty_cols = ttk.Label(sec_e, text="Columnas vacías: —",
            foreground="#8E44AD", font=("TkDefaultFont", 9, "bold"))
        self.lbl_total_empty_cols.pack(anchor=tk.W, pady=1)

        ttk.Separator(sec_e, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)

        self._section_label(sec_e, "Celdas vacías por posición")
        self.auto_cell_text = self._text_box(sec_e, height=5, bg="#FEF9E7")

        self.btn_del_empty_rows = _colored_btn(
            sec_e, "🗑  Eliminar filas con celdas vacías",
            self._delete_rows_with_empty_cells, CLR_RED, CLR_RED_H, tk.DISABLED)
        self.btn_del_empty_rows.pack(fill=tk.X, pady=(4, 8), ipady=2)

        ttk.Separator(sec_e, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=4)

        self._section_label(sec_e, "Columnas completamente vacías")
        self.auto_col_text = self._text_box(sec_e, height=4, bg="#FDFEFE")

        self.btn_del_empty_cols = _colored_btn(
            sec_e, "🗑  Eliminar columnas vacías",
            self._delete_empty_columns, CLR_RED, CLR_RED_H, tk.DISABLED)
        self.btn_del_empty_cols.pack(fill=tk.X, pady=(4, 4), ipady=2)

        # ════════════════════════════════════════════════════════════════════
        # SECCIÓN 2 — DUPLICADOS
        # ════════════════════════════════════════════════════════════════════
        sec_d = ttk.LabelFrame(inner, text="  Registros duplicados  ", padding=10)
        sec_d.pack(fill=tk.X, padx=8, pady=(6, 10))

        self._section_label(sec_d, "Resumen general")
        self.lbl_total_dups = ttk.Label(sec_d, text="Filas duplicadas: —",
            foreground=CLR_ORANGE, font=("TkDefaultFont", 9, "bold"))
        self.lbl_total_dups.pack(anchor=tk.W, pady=1)

        self.lbl_unique_rows = ttk.Label(sec_d, text="Filas únicas: —",
            foreground=CLR_GREEN, font=("TkDefaultFont", 9, "bold"))
        self.lbl_unique_rows.pack(anchor=tk.W, pady=1)

        self.lbl_total_dup_vals = ttk.Label(sec_d, text="Valores duplicados: —",
            foreground=CLR_RED, font=("TkDefaultFont", 9, "bold"))
        self.lbl_total_dup_vals.pack(anchor=tk.W, pady=1)

        ttk.Separator(sec_d, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)

        self._section_label(sec_d, "Posición de filas duplicadas")
        self.auto_dup_text = self._text_box(sec_d, height=4, bg="#FFF8F0")

        self.btn_del_dups = _colored_btn(
            sec_d, "🗑  Eliminar filas duplicadas",
            self._delete_duplicates, CLR_RED, CLR_RED_H, tk.DISABLED)
        self.btn_del_dups.pack(fill=tk.X, pady=(4, 8), ipady=2)

        ttk.Separator(sec_d, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=4)

        self._section_label(sec_d, "Valores duplicados en el archivo")
        self.auto_val_dup_text = self._text_box(sec_d, height=5, bg="#FEF9E7")

        ttk.Separator(sec_d, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)

        self._section_label(sec_d, "Análisis de columnas duplicadas")
        self.auto_col_dup_text = self._text_box(sec_d, height=5, bg="#EAF4FB")

        ttk.Separator(sec_d, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)

        self._section_label(sec_d, "Posición de columnas duplicadas")
        self.auto_col_pos_dup_text = self._text_box(sec_d, height=4, bg="#F0F8FF")

        self.btn_del_dup_cols = _colored_btn(
            sec_d, "🗑  Eliminar columnas duplicadas",
            self._delete_duplicate_columns, CLR_RED, CLR_RED_H, tk.DISABLED)
        self.btn_del_dup_cols.pack(fill=tk.X, pady=(4, 4), ipady=2)

    # ── Utilidad: cuadro de texto con scroll ──────────────────────────────────

    @staticmethod
    def _text_box(parent, height: int, bg: str = "#FFFFFF") -> tk.Text:
        """Crea un tk.Text de solo lectura con scrollbar vertical integrada."""
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, pady=(2, 6))
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL)
        txt = tk.Text(
            frame, height=height, wrap=tk.WORD, state=tk.DISABLED,
            relief=tk.FLAT, bg=bg, font=("TkDefaultFont", 9),
            foreground="#2C3E50", yscrollcommand=vsb.set)
        vsb.config(command=txt.yview)
        txt.grid(row=0, column=0, sticky=tk.NSEW)
        vsb.grid(row=0, column=1, sticky=tk.NS)
        return txt

    # ══════════════════════════════════════════════════════════════════════════
    #  ANÁLISIS COMBINADO (botón "Escanear archivo")
    # ══════════════════════════════════════════════════════════════════════════

    def _scan_all(self) -> None:
        """Ejecuta el análisis de vacíos y duplicados en una sola pasada."""
        if self.dataframe is None:
            self._set_status("Carga un archivo antes de escanear.", "warn")
            return
        self._refresh_auto_panel()
        self._refresh_duplicates()
        self._set_status("Escaneo completado.", "ok")

    # ══════════════════════════════════════════════════════════════════════════
    #  ANÁLISIS DE VACÍOS
    # ══════════════════════════════════════════════════════════════════════════

    def _refresh_auto_panel(self) -> None:
        from openpyxl.utils import get_column_letter

        def _write(widget: tk.Text, msg: str) -> None:
            widget.config(state=tk.NORMAL)
            widget.delete("1.0", tk.END)
            widget.insert(tk.END, msg)
            widget.config(state=tk.DISABLED)

        if self.dataframe is None:
            _write(self.auto_cell_text, "Carga un archivo para analizar.")
            _write(self.auto_col_text,  "Carga un archivo para analizar.")
            self.lbl_total_empty_cells.config(text="Celdas vacías: —")
            self.lbl_total_empty_cols.config(text="Columnas vacías: —")
            self.btn_del_empty_rows.config(state=tk.DISABLED)
            self.btn_del_empty_cols.config(state=tk.DISABLED)
            return

        df   = self.dataframe
        cols = list(df.columns)

        def _is_empty(val) -> bool:
            return pd.isna(val) or (isinstance(val, str) and val.strip() == "")

        empty_cols = [c for c in cols if df[c].apply(_is_empty).all()]

        empty_cells: list[str] = []
        for ci, col in enumerate(cols):
            letter = get_column_letter(ci + 1)
            for ri in df.index:
                if _is_empty(df.at[ri, col]):
                    empty_cells.append(f"{letter}{int(ri) + 2}")

        rows_with_empty = df[
            df.apply(lambda row: row.apply(_is_empty).any(), axis=1)
        ].index.tolist()

        self.lbl_total_empty_cells.config(text=f"Celdas vacías: {len(empty_cells)}")
        self.lbl_total_empty_cols.config(text=f"Columnas vacías: {len(empty_cols)}")
        self.btn_del_empty_rows.config(state=tk.NORMAL if rows_with_empty else tk.DISABLED)
        self.btn_del_empty_cols.config(state=tk.NORMAL if empty_cols      else tk.DISABLED)

        if not empty_cells:
            _write(self.auto_cell_text, "✔ No se encontraron celdas vacías.")
        else:
            from collections import defaultdict
            by_col: dict[str, list[str]] = defaultdict(list)
            for ref in empty_cells:
                by_col["".join(c for c in ref if c.isalpha())].append(ref)
            lines = []
            for letter, refs in by_col.items():
                ci = ord(letter) - ord("A") if len(letter) == 1 \
                     else (ord(letter[0]) - ord("A") + 1) * 26 + (ord(letter[1]) - ord("A"))
                col_name = cols[ci] if ci < len(cols) else letter
                lines.append(
                    f'• "{col_name}": {"celda vacía en" if len(refs)==1 else "celdas vacías en"}'
                    f' {", ".join(refs)}.')
            _write(self.auto_cell_text, "\n".join(lines))

        _write(self.auto_col_text,
               "✔ Todas las columnas contienen al menos un dato." if not empty_cols
               else "\n".join(f'• La columna "{c}" no contiene ningún dato.'
                              for c in empty_cols))

    # ══════════════════════════════════════════════════════════════════════════
    #  ELIMINACIÓN DE VACÍOS
    # ══════════════════════════════════════════════════════════════════════════

    def _delete_empty_columns(self) -> None:
        if self.dataframe is None:
            return
        _e = lambda v: pd.isna(v) or (isinstance(v, str) and v.strip() == "")
        drop = [c for c in self.dataframe.columns if self.dataframe[c].apply(_e).all()]
        if not drop:
            self._set_status("No hay columnas completamente vacías.", "warn"); return
        names = ", ".join(f'"{c}"' for c in drop)
        if not messagebox.askyesno("Confirmar",
                f"Se eliminarán {len(drop)} columna(s) vacía(s):\n{names}\n\n¿Continuar?"):
            return
        self.dataframe.drop(columns=drop, inplace=True)
        self._reset_selection()
        self._display_data(); self._refresh_auto_panel()
        self._set_status(f"{len(drop)} columna(s) vacía(s) eliminadas.", "ok")

    def _delete_rows_with_empty_cells(self) -> None:
        if self.dataframe is None:
            return
        _e = lambda v: pd.isna(v) or (isinstance(v, str) and v.strip() == "")
        mask  = self.dataframe.apply(lambda row: row.apply(_e).any(), axis=1)
        n     = int(mask.sum())
        if n == 0:
            self._set_status("No hay filas con celdas vacías.", "warn"); return
        if not messagebox.askyesno("Confirmar",
                f"Se eliminarán {n} fila(s) con celdas vacías.\n\n¿Continuar?"):
            return
        self.dataframe = self.dataframe[~mask].reset_index(drop=True)
        self._reset_selection()
        self._display_data(); self._refresh_auto_panel()
        self._set_status(f"{n} fila(s) con celdas vacías eliminadas.", "ok")

    # ══════════════════════════════════════════════════════════════════════════
    #  ANÁLISIS DE DUPLICADOS
    # ══════════════════════════════════════════════════════════════════════════

    def _refresh_duplicates(self) -> None:
        def _write(widget: tk.Text, msg: str) -> None:
            widget.config(state=tk.NORMAL)
            widget.delete("1.0", tk.END)
            widget.insert(tk.END, msg)
            widget.config(state=tk.DISABLED)

        if self.dataframe is None:
            for w in (self.auto_dup_text, self.auto_val_dup_text,
                      self.auto_col_dup_text, self.auto_col_pos_dup_text):
                _write(w, "Carga un archivo para analizar.")
            self.lbl_total_dups.config(text="Filas duplicadas: —")
            self.lbl_unique_rows.config(text="Filas únicas: —")
            self.lbl_total_dup_vals.config(text="Valores duplicados: —")
            self.btn_del_dups.config(state=tk.DISABLED)
            self.btn_del_dup_cols.config(state=tk.DISABLED)
            return

        df   = self.dataframe
        cols = list(df.columns)

        # ── NIVEL 1: Filas duplicadas ─────────────────────────────────────────
        dup_mask    = df.duplicated(keep="first")
        dup_indices = df.index[dup_mask].tolist()
        n_dups      = len(dup_indices)
        n_unique    = len(df) - n_dups

        self.lbl_total_dups.config(text=f"Filas duplicadas: {n_dups}")
        self.lbl_unique_rows.config(text=f"Filas únicas: {n_unique}")
        self.btn_del_dups.config(state=tk.NORMAL if n_dups > 0 else tk.DISABLED)

        if n_dups == 0:
            _write(self.auto_dup_text, "✔ No se encontraron filas duplicadas.")
        else:
            MAX = 150
            refs = [f"Fila {int(i)+2}" for i in dup_indices[:MAX]]
            msg  = f"Las siguientes {n_dups} fila(s) son copia de otra anterior:\n"
            msg += ", ".join(refs)
            if n_dups > MAX:
                msg += f"\n… y {n_dups-MAX} más."
            _write(self.auto_dup_text, msg)

        # ── NIVEL 2: Valores duplicados ───────────────────────────────────────
        flat = pd.Series(df.values.flatten()).dropna().astype(str).str.strip()
        flat = flat[flat != ""]
        freq     = flat.value_counts()
        dup_vals = freq[freq > 1]

        self.lbl_total_dup_vals.config(text=f"Valores duplicados: {len(dup_vals)}")

        if dup_vals.empty:
            _write(self.auto_val_dup_text, "✔ No se encontraron valores duplicados.")
        else:
            MAX_V = 60
            lines = [f'• "{v}"  →  aparece {c} veces'
                     for v, c in dup_vals.head(MAX_V).items()]
            msg = "\n".join(lines)
            if len(dup_vals) > MAX_V:
                msg += f"\n… y {len(dup_vals)-MAX_V} valores más."
            _write(self.auto_val_dup_text, msg)

        # ── NIVEL 3: Columnas duplicadas ──────────────────────────────────────
        exact_pairs:   list[tuple[str, str]] = []
        paired: set[str] = set()

        for i in range(len(cols)):
            for j in range(i + 1, len(cols)):
                ca, cb = cols[i], cols[j]
                if (df[ca].astype(str).reset_index(drop=True)
                        .equals(df[cb].astype(str).reset_index(drop=True))):
                    exact_pairs.append((ca, cb))
                    paired.update([ca, cb])

        non_paired = [c for c in cols if c not in paired]
        poss_col, poss_cnt = None, 0
        for col in non_paired:
            n = int(df[col].astype(str).duplicated(keep="first").sum())
            if n > poss_cnt:
                poss_cnt, poss_col = n, col

        # Análisis narrativo
        lines = []
        if not exact_pairs and not poss_col:
            lines.append("✔ No se detectaron columnas duplicadas o sospechosas.")
        else:
            if exact_pairs:
                lines.append("📋 Columnas exactamente duplicadas:")
                for ca, cb in exact_pairs:
                    lines.append(f'  • "{cb}" es copia exacta de "{ca}".')
            if poss_col and poss_cnt > 0:
                pct = round(poss_cnt / len(df) * 100, 1)
                lines += ["",
                          "⚠ Columna posiblemente duplicada:",
                          f'  • "{poss_col}" tiene {poss_cnt} valor(es) repetidos '
                          f'internamente ({pct}% de sus filas).']
        _write(self.auto_col_dup_text, "\n".join(lines))

        # Posición de columnas duplicadas + botón
        self._exact_dup_col_pairs = exact_pairs

        if not exact_pairs:
            _write(self.auto_col_pos_dup_text, "✔ No hay columnas exactamente duplicadas.")
            self.btn_del_dup_cols.config(state=tk.DISABLED)
        else:
            from openpyxl.utils import get_column_letter
            col_list = list(df.columns)
            pos = [f"Se detectaron {len(exact_pairs)} columna(s) duplicadas "
                   "(se conservará la original de cada par):\n"]
            for ca, cb in exact_pairs:
                la = get_column_letter(col_list.index(ca) + 1)
                lb = get_column_letter(col_list.index(cb) + 1)
                pos.append(f'• "{cb}" ({lb})  es copia de  "{ca}" ({la}).')
            _write(self.auto_col_pos_dup_text, "\n".join(pos))
            self.btn_del_dup_cols.config(state=tk.NORMAL)

    # ══════════════════════════════════════════════════════════════════════════
    #  ELIMINACIÓN DE DUPLICADOS
    # ══════════════════════════════════════════════════════════════════════════

    def _delete_duplicates(self) -> None:
        if self.dataframe is None:
            return
        mask  = self.dataframe.duplicated(keep="first")
        n     = int(mask.sum())
        if n == 0:
            self._set_status("No hay filas duplicadas.", "warn"); return
        if not messagebox.askyesno("Confirmar",
                f"Se eliminarán {n} fila(s) duplicadas (se conserva la primera "
                "ocurrencia).\n\n¿Continuar?"):
            return
        self.dataframe = self.dataframe[~mask].reset_index(drop=True)
        self._reset_selection()
        self._display_data(); self._refresh_duplicates(); self._refresh_auto_panel()
        self._set_status(f"{n} fila(s) duplicadas eliminadas.", "ok")

    def _delete_duplicate_columns(self) -> None:
        if self.dataframe is None:
            return
        pairs = self._exact_dup_col_pairs
        if not pairs:
            self._set_status(
                "No hay columnas duplicadas detectadas. Ejecuta 'Escanear' primero.", "warn")
            return
        drop = list(dict.fromkeys(cb for _, cb in pairs))
        names = "\n".join(f'  • "{c}"' for c in drop)
        if not messagebox.askyesno("Confirmar",
                f"Se eliminarán {len(drop)} columna(s) duplicada(s):\n\n{names}"
                "\n\n¿Continuar?"):
            return
        self.dataframe.drop(columns=drop, inplace=True)
        self._exact_dup_col_pairs = []
        self._reset_selection()
        self._display_data(); self._refresh_duplicates(); self._refresh_auto_panel()
        self._set_status(f"{len(drop)} columna(s) duplicada(s) eliminadas.", "ok")

    # ══════════════════════════════════════════════════════════════════════════
    #  UTILIDADES
    # ══════════════════════════════════════════════════════════════════════════

    def _reset_selection(self) -> None:
        """Limpia selección y formatos huérfanos tras operaciones masivas."""
        self.cell_formats    = {}
        self.edited_cells    = {}
        self.selected_cell   = None
        self.selected_column = None

    def _cast_value(self, col_index: int, raw: str):
        col_name = self.dataframe.columns[col_index]
        dtype    = self.dataframe[col_name].dtype
        try:
            if pd.api.types.is_integer_dtype(dtype):
                return int(float(raw.replace(",", ".")))
            elif pd.api.types.is_float_dtype(dtype):
                return float(raw.replace(",", "."))
            elif pd.api.types.is_bool_dtype(dtype):
                return raw.strip().lower() in ("true", "1", "sí", "si", "yes")
            return raw
        except (ValueError, AttributeError):
            return raw

    # ══════════════════════════════════════════════════════════════════════════
    #  CARGA Y VISUALIZACIÓN
    # ══════════════════════════════════════════════════════════════════════════

    def _drop_file(self, event) -> None:
        """Carga un archivo arrastrado a la interfaz."""
        try:
            path = event.data.strip("{}")

            # Validar extensión
            if not path.lower().endswith(".xlsx"):
                self._set_status(
                    "Solo se permiten archivos XLSX.",
                    "warn"
                )
                return

            self.current_file = path
            self.dataframe = pd.read_excel(path)

            self._reset_selection()
            self._display_data()

            self.file_label.config(
                text=f"📄  {os.path.basename(path)}"
            )

            self.save_btn.config(state=tk.NORMAL)

            self._refresh_auto_panel()

            self._set_status(
                f"Archivo '{os.path.basename(path)}' cargado correctamente.",
                "ok"
            )

        except Exception as exc:
            messagebox.showerror(
                "Error al cargar archivo",
                str(exc)
            )

            self._set_status(
                "Error al cargar el archivo arrastrado.",
                "error",
                autoclean=False
            )

    def upload_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Seleccionar archivo XLSX",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        if not path:
            return
        try:
            self.current_file    = path
            self.dataframe       = pd.read_excel(path)
            self._reset_selection()
            self._display_data()
            self.file_label.config(text=f"📄  {os.path.basename(path)}")
            self.save_btn.config(state=tk.NORMAL)
            self._refresh_auto_panel()
            self._set_status(f"Archivo '{os.path.basename(path)}' cargado correctamente.", "ok")
        except Exception as exc:
            messagebox.showerror("Error al cargar", str(exc))
            self._set_status("Error al cargar el archivo.", "error", autoclean=False)

    def _display_data(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.row_id_map = {}

        if self.dataframe is None:
            return

        cols = list(self.dataframe.columns)
        self.tree["columns"] = cols
        self.tree.column("#0", width=50, anchor=tk.CENTER, stretch=False)
        self.tree.heading("#0", text="Fila")

        for i, col in enumerate(cols):
            self.tree.column(col, width=110, anchor=tk.W, minwidth=60)
            self.tree.heading(
                col, text=str(col),
                command=lambda idx=i, name=col: self._on_column_header_click(idx, name)
            )

        for idx, row in self.dataframe.iterrows():
            tag = "evenrow" if int(idx) % 2 == 0 else "oddrow"
            item_id = self.tree.insert("", tk.END, text=str(idx),
                                       values=[str(v) for v in row], tags=(tag,))
            self.row_id_map[item_id] = idx

    # ══════════════════════════════════════════════════════════════════════════
    #  RESOLUCIÓN DE CLICS
    # ══════════════════════════════════════════════════════════════════════════

    def _parse_col_index(self, event) -> int:
        raw = self.tree.identify_column(event.x)
        return int(raw.lstrip("#")) - 1

    def _on_click(self, event) -> None:
        region  = self.tree.identify_region(event.x, event.y)
        if region == "heading":
            return
        if region != "cell":
            return
        item_id = self.tree.identify_row(event.y)
        if not item_id or item_id not in self.row_id_map:
            return
        col_idx = self._parse_col_index(event)
        if col_idx < 0:
            return
        self._select_cell(item_id, col_idx)

    def _on_double_click(self, event) -> None:
        if self.tree.identify_region(event.x, event.y) != "cell":
            return
        item_id = self.tree.identify_row(event.y)
        if not item_id or item_id not in self.row_id_map:
            return
        col_idx = self._parse_col_index(event)
        if col_idx < 0:
            return
        row_index = self.row_id_map[item_id]
        current   = self.tree.item(item_id)["values"][col_idx]
        self._open_edit_window(row_index, col_idx, current)

    def _select_cell(self, item_id: str, col_idx: int) -> None:
        row_index = self.row_id_map[item_id]
        self.selected_cell   = (row_index, col_idx)
        self.selected_column = None

        self._clear_highlight()
        self.tree.item(item_id, tags=("cell_selected",))

        col_name      = self.dataframe.columns[col_idx]
        current_value = self.tree.item(item_id)["values"][col_idx]

        self.cell_info_label.config(text=f"Columna: {col_name}  |  Fila: {row_index}")
        self.value_entry.config(state=tk.NORMAL)
        self.value_entry.delete(0, tk.END)
        self.value_entry.insert(0, str(current_value))

        fmt = self.cell_formats.get((row_index, col_idx), {})
        self.font_var.set(fmt.get("font", "Calibri"))
        self.font_size_var.set(str(fmt.get("size", FONT_SIZE_DEFAULT)))
        self.h_align_var.set(fmt.get("h_align", "left"))
        self.v_align_var.set(fmt.get("v_align", "center"))
        self.number_format_var.set(fmt.get("number_format", "General"))
        self.apply_btn.config(state=tk.NORMAL)

    def _on_column_header_click(self, col_idx: int, col_name: str) -> None:
        self.selected_cell   = None
        self.selected_column = col_idx

        self._clear_highlight()
        for item_id in self.tree.get_children():
            self.tree.item(item_id, tags=("col_selected",))

        self.cell_info_label.config(
            text=f"Columna: {col_name}  ({len(self.dataframe)} filas)")
        self.value_entry.delete(0, tk.END)
        self.value_entry.config(state=tk.DISABLED)

        fmt = self.cell_formats.get((None, col_idx), {})
        self.font_var.set(fmt.get("font", "Calibri"))
        self.font_size_var.set(str(fmt.get("size", FONT_SIZE_DEFAULT)))
        self.h_align_var.set(fmt.get("h_align", "left"))
        self.v_align_var.set(fmt.get("v_align", "center"))
        self.number_format_var.set(fmt.get("number_format", "General"))
        self.apply_btn.config(state=tk.NORMAL)

    def _clear_highlight(self) -> None:
        for item_id in self.tree.get_children():
            self.tree.item(item_id, tags=())

    # ══════════════════════════════════════════════════════════════════════════
    #  EDICIÓN
    # ══════════════════════════════════════════════════════════════════════════

    def _open_edit_window(self, row_index: int, col_index: int,
                          current_value) -> None:
        win = tk.Toplevel(self.root)
        win.title("Editar celda")
        win.geometry("320x120")
        win.configure(bg=BG_BASE)
        win.transient(self.root)
        win.grab_set()
        win.resizable(False, False)

        col_name = self.dataframe.columns[col_index]
        ttk.Label(win, text=f"Columna: {col_name}  —  Fila: {row_index}",
                  background=BG_BASE).pack(pady=8)

        entry = ttk.Entry(win)
        entry.insert(0, str(current_value))
        entry.pack(pady=4, padx=12, fill=tk.X)
        entry.select_range(0, tk.END)
        entry.focus()

        def _save():
            new_val = self._cast_value(col_index, entry.get())
            self.dataframe.at[row_index, col_name] = new_val
            self.edited_cells[(row_index, col_index)] = new_val
            self._display_data()
            self._set_status(f"Valor actualizado en {col_name} / Fila {row_index}.", "ok")
            win.destroy()

        btn_row = tk.Frame(win, bg=BG_BASE)
        btn_row.pack(pady=8)
        _colored_btn(btn_row, "Guardar",   _save,        CLR_GREEN, CLR_GREEN_H).pack(side=tk.LEFT, padx=5)
        _colored_btn(btn_row, "Cancelar",  win.destroy,  CLR_RED,   CLR_RED_H  ).pack(side=tk.LEFT, padx=5)
        entry.bind("<Return>", lambda _: _save())
        entry.bind("<Escape>", lambda _: win.destroy())

    def _apply_formatting(self) -> None:
        try:
            size = int(self.font_size_var.get())
            if not (FONT_SIZE_MIN <= size <= FONT_SIZE_MAX):
                raise ValueError
        except ValueError:
            messagebox.showerror("Error",
                f"El tamaño debe estar entre {FONT_SIZE_MIN} y {FONT_SIZE_MAX}.")
            return

        fmt = {
            "font":          self.font_var.get(),
            "size":          size,
            "h_align":       self.h_align_var.get(),
            "v_align":       self.v_align_var.get(),
            "number_format": self.number_format_var.get(),
        }

        if self.selected_cell is not None:
            row_index, col_index = self.selected_cell
            self.cell_formats[(row_index, col_index)] = fmt
            col_name  = self.dataframe.columns[col_index]
            new_value = self._cast_value(col_index, self.value_entry.get())
            self.dataframe.at[row_index, col_name] = new_value
            self.edited_cells[(row_index, col_index)] = new_value
            self._display_data()
            self._set_status("Formato aplicado a la celda.", "ok")

        elif self.selected_column is not None:
            col_index = self.selected_column
            for idx in self.dataframe.index:
                self.cell_formats[(idx, col_index)] = fmt
            self.cell_formats[(None, col_index)] = fmt
            col_name = self.dataframe.columns[col_index]
            self._display_data()
            self._set_status(f"Formato aplicado a la columna '{col_name}'.", "ok")

        else:
            self._set_status("Selecciona una celda o columna primero.", "warn")

    # ══════════════════════════════════════════════════════════════════════════
    #  GUARDADO
    # ══════════════════════════════════════════════════════════════════════════

    def save_file(self) -> None:
        if self.current_file is None or self.dataframe is None:
            self._set_status("No hay archivo cargado.", "warn"); return
        try:
            self.dataframe.to_excel(self.current_file, index=False)
            wb = load_workbook(self.current_file)
            ws = wb.active

            for (row_index, col_index), fmt in self.cell_formats.items():
                if row_index is None:
                    continue
                cell = ws.cell(row=row_index + 2, column=col_index + 1)
                cell.font = Font(
                    name=fmt.get("font", "Calibri"),
                    size=fmt.get("size", FONT_SIZE_DEFAULT))
                h = fmt.get("h_align", "left")
                v = fmt.get("v_align", "center")
                cell.alignment = Alignment(
                    horizontal=h if h in ALIGN_OPTIONS  else "left",
                    vertical=  v if v in VALIGN_OPTIONS else "center",
                    wrap_text=True)
                cell.number_format = NUMBER_FORMAT_MAP.get(
                    fmt.get("number_format", "General"), "General")

            wb.save(self.current_file)
            self.edited_cells = {}
            self.cell_formats = {}
            self._set_status(
                f"Archivo guardado correctamente: {os.path.basename(self.current_file)}", "ok")
        except Exception as exc:
            messagebox.showerror("Error al guardar", str(exc))
            self._set_status("Error al guardar el archivo.", "error", autoclean=False)


# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app  = DataCleanerApp(root)
    root.mainloop()